import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt5 import uic
from PyQt5.QtCore import Qt, QEvent
from veritabani import Veritabani
from PyQt5.QtWidgets import QTableWidgetItem
from PyQt5.QtWidgets import QHeaderView
from PyQt5.QtWidgets import QAbstractItemView
from PyQt5.QtWidgets import QPushButton
from datetime import datetime
import openpyxl
from PyQt5.QtWidgets import QFileDialog


class AnaSayfa(QMainWindow):
    
    def __init__(self, kullanici_bilgisi):
        super().__init__()
        uic.loadUi('anasayfa.ui', self)
        
        header = self.tbl_uyeler.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)

        self.tbl_uyeler.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_uyeler.setSelectionMode(QAbstractItemView.SingleSelection)
        
        self.db = Veritabani()
        self.kullanici = kullanici_bilgisi
        self.secili_odunc_uye_id = None
        self.secili_odunc_kitap_id = None
        
        if isinstance(kullanici_bilgisi, (tuple, list)):
             self.kullanici_id = kullanici_bilgisi[0] 
        else:
             self.kullanici_id = 1

        self.kategori_combo_doldur()
        self.kitap_listele()
        self.uye_listele()
        self.odunc_uye_listele()
        self.odunc_kitap_listele()
        self.aktif_oduncleri_listele()
        
        ad_soyad = self.kullanici['AdSoyad']
        rol = self.kullanici['Rol']
        self.lbl_baslik.setText(f"Hos geldiniz: {ad_soyad} ({rol})")
        self.lbl_secili_uye_id.setText("-")

        self.ui_baglantilari()
        
        self.teslim_listesi_yukle()
        self.lne_teslim_ara.textChanged.connect(self.teslim_listesi_yukle)             
        self.tbl_teslim_listesi.itemSelectionChanged.connect(self.teslim_satir_secildi)
        self.stackedWidget.currentChanged.connect(self.sayfa_kontrolu)
            
            
            
        self.tbl_teslim_listesi.clicked.connect(self.teslim_satir_secildi)
            
       
        
        # --- CEZA EKRANI BAĞLANTILARI ---
        self.btn_ceza_filtrele.clicked.connect(self.cezalari_listele)
        
        self.ceza_ekrani_hazirlik()
        
        
        # --- RAPOR EKRANI BAĞLANTILARI ---
        self.btn_rapor_filtrele.clicked.connect(self.rapor_hareket_listele)
            
        self.btn_rapor_geciken_yenile.clicked.connect(self.rapor_geciken_listele)
            
        self.btn_rapor_populer_yenile.clicked.connect(self.rapor_populer_listele)

        self.rapor_ekrani_hazirlik()
        self.btn_rapor_aktif_yenile.clicked.connect(self.rapor_aktif_listele)
            
        self.btn_rapor_ceza_yenile.clicked.connect(self.rapor_ceza_listele)
            
        self.btn_sorgu_bul.clicked.connect(self.dinamik_sorgu_calistir)
            
        self.btn_sorgu_excel.clicked.connect(self.sonuclari_excele_aktar)
        self.btn_sorgu_temizle.clicked.connect(self.filtreleri_temizle)
        self.sorgu_ekrani_hazirlik()    
        self.btn_sorgu.clicked.connect(self.sorgu_sayfasini_ac)
        
        
    def sayfa_degistir(self, index):
        self.stackedWidget.setCurrentIndex(index)

    def cikis_yap(self):
        answer = QMessageBox.question(self, "Çıkış", "Çıkmak istediğinize emin misiniz?", 
                                      QMessageBox.Yes | QMessageBox.No)
        if answer == QMessageBox.Yes:
            self.close()
    def ui_baglantilari(self):
        try: self.lne_uye_ara.textChanged.disconnect()
        except: pass
        
        # --- Üye İşlemleri ---
        self.btn_uye_ekle.clicked.connect(self.uye_ekle)
        self.btn_uye_sil.clicked.connect(self.uye_sil)
        self.btn_uye_guncelle.clicked.connect(self.uye_guncelle_islemi)
        self.lne_uye_ara.textChanged.connect(self.uye_ara)
        self.btn_temizle.clicked.connect(self.formu_temizle)
        self.tbl_uyeler.cellClicked.connect(self.satir_secildi)

        # --- Kitap İşlemleri ---
        self.btn_kitap_ekle.clicked.connect(self.kitap_ekle)
        self.btn_kitap_sil.clicked.connect(self.kitap_sil)
        self.btn_kitap_guncelle.clicked.connect(self.kitap_guncelle)
        self.lne_kitap_ara.textChanged.connect(self.kitap_ara)
        self.tbl_kitaplar.cellClicked.connect(self.kitap_secildi)
        self.btn_kitap_temizle.clicked.connect(self.kitap_formu_temizle)
        self.btn_kitap_teslim_al.clicked.connect(self.kitap_teslim_al)
        # --- Kategori İşlemleri ---
        self.btn_kat_ekle.clicked.connect(self.yeni_kategori_ekle)
        self.btn_kat_sil.clicked.connect(self.secili_kategoriyi_sil)
        self.btn_kat_guncelle.clicked.connect(self.secili_kategoriyi_guncelle)

        # --- Ödünç Verme İşlemleri ---
        self.lne_odunc_uye_ara.textChanged.connect(self.odunc_uye_listele)
        self.lne_odunc_kitap_ara.textChanged.connect(self.odunc_kitap_listele)
        self.tbl_odunc_uyeler.clicked.connect(self.odunc_uye_secildi)
        self.tbl_odunc_kitaplar.clicked.connect(self.odunc_kitap_secildi)
        self.btn_odunc_ver.clicked.connect(self.odunc_ver_islemi)
        try:
            self.btn_filtre_temizle.clicked.connect(self.odunc_filtre_temizle)
        except: pass

        # --- Sayfa Navigasyonu (Sol Menü) ---
        self.btn_uye.clicked.connect(lambda: self.sayfa_degistir(0))   
        self.btn_kitap.clicked.connect(lambda: self.sayfa_degistir(1)) 
        self.btn_odunc.clicked.connect(lambda: self.sayfa_degistir(2)) 
        self.btn_menu_teslim_al.clicked.connect(lambda: self.sayfa_degistir(3)) 
        self.btn_ceza.clicked.connect(lambda: self.sayfa_degistir(4)) 
        self.btn_rapor.clicked.connect(lambda: self.sayfa_degistir(5)) 
        self.btn_sorgu.clicked.connect(lambda: self.sayfa_degistir(6)) 
        self.stackedWidget.currentChanged.connect(self.sayfa_kontrolu)
        self.btn_cikis.clicked.connect(self.cikis_yap)
    
            

    def uye_ekle(self):
        ad = self.lne_uye_ad.text()
        soyad = self.lne_uye_soyad.text()
        email = self.lne_uye_email.text()
        tel = self.lne_uye_tel.text()

        if ad and soyad and email:
            basarili = self.db.uye_ekle(ad, soyad, email, tel)
            if basarili:
                QMessageBox.information(self, "Başarılı", "Üye başarıyla eklendi.")
                self.uye_listele()
                self.kutulari_temizle()
            else:
                QMessageBox.warning(self, "Hata", "Ekleme başarısız oldu.")
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen zorunlu alanları doldurun!")

    def satir_secildi(self):
        secili_satir = self.tbl_uyeler.currentRow()
        if secili_satir < 0: return

        ek_bilgi = "" 
        try:
            uye_id = self.tbl_uyeler.item(secili_satir, 0).text()
            
            ozet = self.db.uye_ozet_bilgi_getir(uye_id)
            
            if ozet:
                
                borc = ozet[1]
                elindeki = ozet[3]
                
                ek_bilgi = f" | 📘 Elindeki kitap sayısı: {elindeki} | 💰 Borç: {borc} TL"
            
            ad = self.tbl_uyeler.item(secili_satir, 1).text()
            soyad = self.tbl_uyeler.item(secili_satir, 2).text()
            
            self.lbl_secili_uye_id.setText(f"Seçilen: {ad} {soyad} (ID: {uye_id}){ek_bilgi}")
            
            self.lne_uye_ad.setText(ad)
            self.lne_uye_soyad.setText(soyad)
            self.lne_uye_email.setText(self.tbl_uyeler.item(secili_satir, 3).text())
            self.lne_uye_tel.setText(self.tbl_uyeler.item(secili_satir, 4).text())
            
            self.secilen_uye_id = uye_id
            
        except Exception as e:
            print(f"Hata: {e}")
            
    def formu_temizle(self):
        self.lne_uye_ad.clear()
        self.lne_uye_soyad.clear()
        self.lne_uye_email.clear()
        self.lne_uye_tel.clear()
        self.lbl_secili_uye_id.setText("Seçilen Üye: Henüz Seçilmedi")
        self.tbl_uyeler.clearSelection()

    def uye_guncelle_islemi(self):
        row = self.tbl_uyeler.currentRow()
        if row < 0:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Hata", "Lütfen listeden güncellenecek bir üye seçin!")
            return

        try:
            secili_id = self.tbl_uyeler.item(row, 0).text()
            
            yeni_ad = self.lne_uye_ad.text()
            yeni_soyad = self.lne_uye_soyad.text()
            yeni_mail = self.lne_uye_email.text()
            yeni_tel = self.lne_uye_tel.text()
            
            basarili = self.db.uye_guncelle(secili_id, yeni_ad, yeni_soyad, yeni_mail, yeni_tel)
            
            if basarili:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.information(self, "Başarılı", "Üye bilgileri güncellendi.")
                
                self.uye_listele()
                
                self.lne_uye_ad.clear()
                self.lne_uye_soyad.clear()
                self.lne_uye_email.clear()
                self.lne_uye_tel.clear()
            else:
                QMessageBox.critical(self, "Hata", "Güncelleme yapılamadı!")
                
        except Exception as e:
            print(f"Hata detayı: {e}")
    def uye_sil(self):
        
        if not hasattr(self, 'secilen_uye_id') or not self.secilen_uye_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen tablodan silinecek bir üye seçin.")
            return

        uye_id = self.secilen_uye_id

        cevap = QMessageBox.question(self, "Sil", "Bu üyeyi silmek istediğinize emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        
        if cevap == QMessageBox.Yes:
            durum, mesaj = self.db.uye_sil(uye_id)
            
            if durum: 
                QMessageBox.information(self, "Başarılı", mesaj)
                self.uye_listele()      
                self.kutulari_temizle() 
                self.secilen_uye_id = None 
            else: 
                QMessageBox.critical(self, "Hata", mesaj)

   
    # --- LİSTELEME ---
    def uye_listele(self):
        uyeler = self.db.uye_listele()
        self.tabloyu_doldur(uyeler)

   
    def uye_ara(self):
        kelime = self.lne_uye_ara.text().strip()

        if kelime == "":
            self.uye_listele()
        else:
            sonuclar = self.db.uye_ara1(kelime) 
            self.tabloyu_doldur(sonuclar)
            
            
    def tabloyu_doldur(self, veri_listesi):
        self.tbl_uyeler.clearContents()
        self.tbl_uyeler.setRowCount(0)
        
        if not veri_listesi:
            return

        for satir_indeks, satir_veri in enumerate(veri_listesi):
            self.tbl_uyeler.insertRow(satir_indeks)
            self.tbl_uyeler.setItem(satir_indeks, 0, QTableWidgetItem(str(satir_veri[0]))) # ID
            self.tbl_uyeler.setItem(satir_indeks, 1, QTableWidgetItem(str(satir_veri[1]))) # Ad
            self.tbl_uyeler.setItem(satir_indeks, 2, QTableWidgetItem(str(satir_veri[2]))) # Soyad
            self.tbl_uyeler.setItem(satir_indeks, 3, QTableWidgetItem(str(satir_veri[3]))) # Email
            self.tbl_uyeler.setItem(satir_indeks, 4, QTableWidgetItem(str(satir_veri[4]))) # Telefon
            self.tbl_uyeler.setItem(satir_indeks, 5, QTableWidgetItem(f"{satir_veri[5]} TL")) # Borç    
    def kutulari_temizle(self):
        self.lne_uye_ad.clear()
        self.lne_uye_soyad.clear()
        self.lne_uye_email.clear()
        self.lne_uye_tel.clear()
        self.lbl_secili_uye_id.clear()
        
    def klavye_navigasyonu_ayarla(self):
        self.setTabOrder(self.lne_uye_ara, self.tbl_uyeler)
        self.setTabOrder(self.tbl_uyeler, self.lne_uye_ad)
        self.setTabOrder(self.lne_uye_ad, self.lne_uye_soyad)
        self.setTabOrder(self.lne_uye_soyad, self.lne_uye_tel)
        self.setTabOrder(self.lne_uye_tel, self.lne_uye_email)
        self.setTabOrder(self.lne_uye_email, self.btn_temizle)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            odaklanilan = self.focusWidget()
            if isinstance(odaklanilan, QPushButton):
                odaklanilan.click()
            else:
                self.focusNextChild()
                
        elif event.key() in (Qt.Key_Down, Qt.Key_Right):
            self.focusNextChild()
        elif event.key() in (Qt.Key_Up, Qt.Key_Left):
            self.focusPreviousChild()
        else:
            super().keyPressEvent(event)
            
            
    #           KİTAP SAYFASI FONKSİYONLARI

    def sayfa_kontrolu(self, index):
        if index == 1:
            self.kitap_sayfasini_yukle()
        elif index == 2:
            self.teslim_listesi_yukle()

    def kitap_sayfasini_yukle(self):
        self.cmb_kategori.clear()
        kategoriler = self.db.kategorileri_getir()
        
        for k in kategoriler:
            if isinstance(k, dict):
                k_id = k['ID']
                k_ad = k['Ad']
            else:
                k_id = k[0]
                k_ad = k[1]
                
            self.cmb_kategori.addItem(k_ad, k_id)
            
        self.kitap_listele()

    def kitap_ekle(self):
        ad = self.lne_kitap_ad.text().strip()
        yazar = self.lne_kitap_yazar.text().strip()
        yayinevi = self.lne_kitap_yayinevi.text().strip()
        yil = self.lne_kitap_yil.text().strip()
        adet_text = self.lne_kitap_adet.text().strip()
        
        kategori_id = self.cmb_kategori.currentData()
        
        if not (ad and yazar and yayinevi and yil and adet_text and kategori_id):
            QMessageBox.warning(self, "Eksik Bilgi", "Lütfen kategori dahil tüm alanları doldurun.")
            return

        try:
            adet = int(adet_text)
        except ValueError:
             QMessageBox.warning(self, "Hata", "Adet ve Yıl kısmına sadece sayı giriniz!")
             return

        basari = self.db.kitap_ekle(ad, yazar, yayinevi, yil, kategori_id, adet)
        
        if basari:
            QMessageBox.information(self, "Başarılı", "Kitap işlemi başarıyla tamamlandı (Eklendi veya Stok Güncellendi).")
            self.kitap_listele() 
            
        else:
            QMessageBox.critical(self, "Hata", "Veritabanı işlem hatası oluştu.")
    def kitap_listele(self):
        self.tbl_kitaplar.clearContents()
        self.tbl_kitaplar.setRowCount(0)
        
        if hasattr(self, 'lne_kitap_ara'):
            arama_metni = self.lne_kitap_ara.text().strip()
        else:
            arama_metni = ""
        
        if arama_metni == "":
            kitaplar = self.db.kitaplari_getir()
        else:
            kitaplar = self.db.kitap_ara(arama_metni, "Genel")
            
        if kitaplar:
            self.tbl_kitaplar.setRowCount(len(kitaplar)) 
            
            for i, k in enumerate(kitaplar):
                
                self.tbl_kitaplar.setItem(i, 0, QTableWidgetItem(str(k[0])))   # ID
                self.tbl_kitaplar.setItem(i, 1, QTableWidgetItem(str(k[1])))   # Kitap Adı
                self.tbl_kitaplar.setItem(i, 2, QTableWidgetItem(str(k[2])))   # Yazar
                self.tbl_kitaplar.setItem(i, 3, QTableWidgetItem(str(k[3])))   # Kategori Adı 
                self.tbl_kitaplar.setItem(i, 4, QTableWidgetItem(str(k[4])))   # Yayınevi 
                self.tbl_kitaplar.setItem(i, 5, QTableWidgetItem(str(k[5])))   # Yıl
                self.tbl_kitaplar.setItem(i, 6, QTableWidgetItem(str(k[6])))   # Stok (Toplam)
                self.tbl_kitaplar.setItem(i, 7, QTableWidgetItem(str(k[7])))   # Mevcut 
      
    def kitap_formu_temizle(self):
        self.lne_kitap_ad.clear()
        self.lne_kitap_yazar.clear()
        self.lne_kitap_yayinevi.clear()
        self.lne_kitap_adet.clear()
        self.lne_kitap_yil.clear()
        self.cmb_kategori.setCurrentIndex(-1)
        self.tbl_kitaplar.clearSelection()
        self.lne_kitap_ad.setFocus()
        
    def kitap_sil(self):
        secili = self.tbl_kitaplar.currentRow()
        if secili < 0:
            QMessageBox.warning(self, "Uyarı", "Silinecek kitabı seçin.")
            return
            
        kitap_id = self.tbl_kitaplar.item(secili, 0).text()
        
        onay = QMessageBox.question(self, "Onay", "Kitabı silmek istediğinize emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        if onay == QMessageBox.Yes:
            if self.db.kitap_sil(kitap_id):
                self.kitap_listele()
            else:
                QMessageBox.critical(self, "Hata", "Kitap silinemedi ödünçte olabilir).")       
    def kitap_secildi(self):
        secili = self.tbl_kitaplar.currentRow()
        if secili < 0: return
        
        self.lne_kitap_ad.setText(self.tbl_kitaplar.item(secili, 1).text())
        self.lne_kitap_yazar.setText(self.tbl_kitaplar.item(secili, 2).text())
        
        kategori_adi = self.tbl_kitaplar.item(secili, 3).text()
        index = self.cmb_kategori.findText(kategori_adi)
        if index >= 0:
            self.cmb_kategori.setCurrentIndex(index) 
            
        self.lne_kitap_yayinevi.setText(self.tbl_kitaplar.item(secili, 4).text())
        self.lne_kitap_yil.setText(self.tbl_kitaplar.item(secili, 5).text())
        self.lne_kitap_adet.setText(self.tbl_kitaplar.item(secili, 6).text())

    def kitap_guncelle(self):
        secili = self.tbl_kitaplar.currentRow()
        if secili < 0:
            QMessageBox.warning(self, "Uyarı", "Güncellenecek kitabı listeden seçin.")
            return

        kitap_id = self.tbl_kitaplar.item(secili, 0).text()

        ad = self.lne_kitap_ad.text().strip()
        yazar = self.lne_kitap_yazar.text().strip()
        yayinevi = self.lne_kitap_yayinevi.text().strip()
        yil = self.lne_kitap_yil.text().strip()
        stok = self.lne_kitap_adet.text().strip()
        kategori_id = self.cmb_kategori.currentData() 
        if ad and yazar and stok:
            basari = self.db.kitap_guncelle(kitap_id, ad, yazar, yayinevi, yil, kategori_id, stok)
            if basari:
                QMessageBox.information(self, "Başarılı", "Kitap bilgileri güncellendi.")
                self.kitap_listele() 
            else:
                QMessageBox.warning(self, "Hata", "Güncelleme yapılamadı.")
        else:
            QMessageBox.warning(self, "Eksik", "Lütfen gerekli alanları doldurun.")

    def kitap_ara(self):
        self.kitap_listele()
            
    def kategori_combo_doldur(self):
        self.cmb_kategori.clear()
        self.cmb_kategori.setPlaceholderText("Kategori Seçiniz...")
        kategoriler = self.db.kategorileri_getir()
        
        for k in kategoriler:
            if isinstance(k, dict):
                ad = k['Ad']
                kid = k['ID']
            else:
                kid = k[0]
                ad = k[1]
                
            self.cmb_kategori.addItem(ad, kid)
        self.cmb_kategori.setCurrentIndex(-1)

    def yeni_kategori_ekle(self):
        yeni_ad = self.lne_yeni_kategori.text().strip()
        
        if yeni_ad:
            basari = self.db.kategori_ekle(yeni_ad)
            if basari:
                QMessageBox.information(self, "Başarılı", "Yeni kategori eklendi.")
                self.lne_yeni_kategori.clear()
                
                self.kategori_combo_doldur()
                
                index = self.cmb_kategori.findText(yeni_ad)
                if index >= 0:
                    self.cmb_kategori.setCurrentIndex(index)
            else:
                QMessageBox.warning(self, "Hata", "Kategori eklenemedi.")
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen kategori adı yazın.")
    def secili_kategoriyi_guncelle(self):
        secili_index = self.cmb_kategori.currentIndex()
        if secili_index < 0:
            QMessageBox.warning(self, "Uyarı", "Güncellenecek kategoriyi soldaki listeden seçin.")
            return

        kat_id = self.cmb_kategori.itemData(secili_index)
        eski_ad = self.cmb_kategori.currentText()
        
        yeni_ad = self.lne_yeni_kategori.text().strip()
        
        if not yeni_ad:
            QMessageBox.warning(self, "Uyarı", "Lütfen kutuya yeni kategori adını yazın.")
            return
            
        soru = f"'{eski_ad}' kategorisinin adını '{yeni_ad}' olarak değiştirmek istiyor musunuz?"
        cevap = QMessageBox.question(self, "Güncelle", soru, QMessageBox.Yes | QMessageBox.No)
        
        if cevap == QMessageBox.Yes:
            durum, mesaj = self.db.kategori_guncelle(kat_id, yeni_ad)
            
            if durum:
                QMessageBox.information(self, "Başarılı", mesaj)
                self.lne_yeni_kategori.clear()
                self.kategori_combo_doldur()
                yeni_index = self.cmb_kategori.findText(yeni_ad.title())
                if yeni_index >= 0:
                    self.cmb_kategori.setCurrentIndex(yeni_index)
            else:
                QMessageBox.critical(self, "Hata", mesaj)        

    def secili_kategoriyi_sil(self):
        secili_index = self.cmb_kategori.currentIndex()
        
        if secili_index < 0:
            QMessageBox.warning(self, "Uyarı", "Silinecek kategori yok veya seçilmedi.")
            return
            
        secili_id = self.cmb_kategori.itemData(secili_index)
        secili_ad = self.cmb_kategori.currentText()
        
        onay = QMessageBox.question(self, "Onay", f"'{secili_ad}' kategorisini silmek istiyor musunuz?", QMessageBox.Yes | QMessageBox.No)
        
        if onay == QMessageBox.Yes:
            durum, mesaj = self.db.kategori_sil(secili_id)
            if durum:
                QMessageBox.information(self, "Başarılı", mesaj)
                self.kategori_combo_doldur()
            else:  
                QMessageBox.critical(self, "Hata", mesaj)
                
    # --- ÖDÜNÇ EKRANI FONKSİYONLARI ---

    def odunc_uye_listele(self):
        ara = self.lne_odunc_uye_ara.text().strip()
        uyeler = self.db.uye_ara(ara) if ara else self.db.uyeleri_getir()
        
        self.tbl_odunc_uyeler.setRowCount(0)
        if uyeler:
            self.tbl_odunc_uyeler.setRowCount(len(uyeler))
            self.tbl_odunc_uyeler.setHorizontalHeaderLabels(["ID", "Ad", "Soyad", "Tel", "Borç"])
            
            for i, u in enumerate(uyeler):
                self.tbl_odunc_uyeler.setItem(i, 0, QTableWidgetItem(str(u[0])))
                self.tbl_odunc_uyeler.setItem(i, 1, QTableWidgetItem(u[1]))
                self.tbl_odunc_uyeler.setItem(i, 2, QTableWidgetItem(u[2]))
                self.tbl_odunc_uyeler.setItem(i, 3, QTableWidgetItem(u[4])) 
                self.tbl_odunc_uyeler.setItem(i, 4, QTableWidgetItem(str(u[5]))) 

    def odunc_kitap_listele(self):
        ara = self.lne_odunc_kitap_ara.text().strip()
        kitaplar = self.db.kitap_ara(ara) if ara else self.db.kitaplari_getir()
        
        self.tbl_odunc_kitaplar.setRowCount(0)
        if kitaplar:
            self.tbl_odunc_kitaplar.setRowCount(len(kitaplar))
            self.tbl_odunc_kitaplar.setHorizontalHeaderLabels(["ID", "Kitap Adı", "Yazar", "Mevcut"])
            
            for i, k in enumerate(kitaplar):
                # k yapısı: ID(0), Ad(1), Yazar(2), Yayinevi(3), Yil(4), Kat(5), Toplam(6), Mevcut(7)
                self.tbl_odunc_kitaplar.setItem(i, 0, QTableWidgetItem(str(k[0])))
                self.tbl_odunc_kitaplar.setItem(i, 1, QTableWidgetItem(k[1]))
                self.tbl_odunc_kitaplar.setItem(i, 2, QTableWidgetItem(k[2]))
                self.tbl_odunc_kitaplar.setItem(i, 3, QTableWidgetItem(str(k[7]))) 
    def odunc_uye_secildi(self):
        secili = self.tbl_odunc_uyeler.currentRow()
        if secili < 0: return 
        
        try:
            id_item = self.tbl_odunc_uyeler.item(secili, 0)
            if id_item:
                self.secili_odunc_uye_id = int(id_item.text())
                
                isim = self.tbl_odunc_uyeler.item(secili, 1).text()
                soyisim = self.tbl_odunc_uyeler.item(secili, 2).text()
                
                self.lbl_secili_uye.setText(f"Seçili Üye: {isim} {soyisim}")
                self.lbl_secili_uye.setStyleSheet("color: green; font-weight: bold;")
                
                self.aktif_oduncleri_listele(self.secili_odunc_uye_id)  
                
        except Exception as e:
            print(f"Üye Seçim Hatası: {e}")
            
    def odunc_filtre_temizle(self):
        self.secili_odunc_uye_id = None
        self.tbl_odunc_uyeler.clearSelection()
        self.lbl_secili_uye.setText("Seçili Üye: Yok")
        self.lbl_secili_uye.setStyleSheet("color: black;") 
        self.aktif_oduncleri_listele()      
            
    def odunc_kitap_secildi(self):
        row = self.tbl_odunc_kitaplar.currentRow()
        if row >= 0:
            self.secili_odunc_kitap_id = int(self.tbl_odunc_kitaplar.item(row, 0).text())
            ad = self.tbl_odunc_kitaplar.item(row, 1).text()
            stok = self.tbl_odunc_kitaplar.item(row, 3).text() 
            
            self.lbl_secili_kitap.setText(f"Seçili Kitap: {ad} (Stok: {stok})")
            self.lbl_secili_kitap.setStyleSheet("color: green; font-weight: bold;")
            

    def aktif_oduncleri_listele(self, uye_id=None):
        self.tbl_uye_aktif_oduncler.clearContents() 
        self.tbl_uye_aktif_oduncler.setRowCount(0)
        
       
        if uye_id:
            sonuclar = self.db.uyenin_aktif_oduncleri(uye_id)
        else:
            sonuclar = self.db.tum_aktif_oduncleri_getir()

        if not sonuclar:
            return
        
        self.tbl_uye_aktif_oduncler.setRowCount(len(sonuclar))
        
        for satir_index, veri in enumerate(sonuclar):
            
            oid = veri[0]
            kitap = veri[1]
            uye_tam_ad = f"{veri[2]} {veri[3]}"
            t1 = veri[4].strftime('%d.%m.%Y') if veri[4] else "..."
            t2 = veri[5].strftime('%d.%m.%Y') if veri[5] else "..."
            
            tarihler = f"{t1} -> {t2}"
            self.tbl_uye_aktif_oduncler.setItem(satir_index, 0, QTableWidgetItem(str(oid)))
            self.tbl_uye_aktif_oduncler.setItem(satir_index, 1, QTableWidgetItem(str(kitap)))
            self.tbl_uye_aktif_oduncler.setItem(satir_index, 2, QTableWidgetItem(str(uye_tam_ad)))
            self.tbl_uye_aktif_oduncler.setItem(satir_index, 3, QTableWidgetItem(str(tarihler)))
        
    def odunc_ver_islemi(self):
        if not self.secili_odunc_uye_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen sol listeden bir üye seçin.")
            return
        if not self.secili_odunc_kitap_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen sağ listeden bir kitap seçin.")
            return
            
        cevap = QMessageBox.question(self, "Onay", "Seçili kitabı bu üyeye ödünç vermek istiyor musunuz?", QMessageBox.Yes | QMessageBox.No)
        
        if cevap == QMessageBox.Yes:
            basarili, mesaj = self.db.odunc_ver(self.secili_odunc_uye_id, self.secili_odunc_kitap_id, self.kullanici_id)
            
            if basarili:
                QMessageBox.information(self, "Başarılı", mesaj)
                self.kitap_listele()
                self.teslim_listesi_yukle()
                self.odunc_kitap_listele()
                self.aktif_oduncleri_listele(self.secili_odunc_uye_id) 
                if self.secili_odunc_uye_id:
                    self.aktif_oduncleri_listele(self.secili_odunc_uye_id)
            else:
                QMessageBox.critical(self, "Hata", mesaj)
                
    # --- LİSTELEME FONKSİYONU ---
    def teslim_listesi_yukle(self):
        
        try:
            aranan = ""
            aranan = self.lne_teslim_ara.text().strip()
            
            sonuclar = self.db.aktif_oduncleri_getir(aranan)

            self.tbl_teslim_listesi.setRowCount(0) # Tabloyu temizle
            self.secili_teslim_id = None 
            self.lbl_teslim_uye.setText("Seçim Bekleniyor...")
            self.lbl_teslim_kitap.setText("-")
            self.lbl_teslim_tarih_verilis.setText("-")
            self.lbl_teslim_tarih_son.setText("-")

            
            for satir_num, veri in enumerate(sonuclar):
                self.tbl_teslim_listesi.insertRow(satir_num)
                
                try:
                    self.tbl_teslim_listesi.setItem(satir_num, 0, QTableWidgetItem(str(veri[0]))) 
                    self.tbl_teslim_listesi.setItem(satir_num, 1, QTableWidgetItem(str(veri[1]))) 
                    self.tbl_teslim_listesi.setItem(satir_num, 2, QTableWidgetItem(str(veri[2]))) 
                    self.tbl_teslim_listesi.setItem(satir_num, 3, QTableWidgetItem(str(veri[3]))) 
                    raw_t1 = veri[4]
                    if hasattr(raw_t1, 'strftime'): # Eğer Tarih Objesiyse
                        tarih1 = raw_t1.strftime('%d.%m.%Y')
                    else: # Eğer Yazıysa (String) - Hata veren yer burasıydı
                        # Yazıyı boşluktan bölüp saati atıyoruz
                        tarih1 = str(raw_t1).split(' ')[0] 

                    self.tbl_teslim_listesi.setItem(satir_num, 4, QTableWidgetItem(tarih1)) 

                    # 4. Sütun: Son Teslim Tarihi
                    raw_t2 = veri[5]
                    if hasattr(raw_t2, 'strftime'):
                        tarih2 = raw_t2.strftime('%d.%m.%Y')
                    else:
                        tarih2 = str(raw_t2).split(' ')[0]

                    self.tbl_teslim_listesi.setItem(satir_num, 5, QTableWidgetItem(tarih2))
                  
                  
                except IndexError as hata:
                    print(f"SÜTUN HATASI: {hata}")
                
                
        except Exception as e:
            print(f"BÜYÜK HATA OLUŞTU: {e}")

    def teslim_satir_secildi(self):
        secili = self.tbl_teslim_listesi.currentRow()
        
        if secili > -1:
            oid = self.tbl_teslim_listesi.item(secili, 0).text()
            ad = self.tbl_teslim_listesi.item(secili, 1).text()
            soyad = self.tbl_teslim_listesi.item(secili, 2).text()
            kitap = self.tbl_teslim_listesi.item(secili, 3).text()
            t_odunc = self.tbl_teslim_listesi.item(secili, 4).text()
            t_son = self.tbl_teslim_listesi.item(secili, 5).text()
            
            self.secili_odunc_id = int(oid)
            
            self.lbl_teslim_uye.setText(f"👤 Üye: {ad} {soyad}")
            self.lbl_teslim_kitap.setText(f"📘 Kitap: {kitap}")
            self.lbl_teslim_tarih_verilis.setText(f"📅 Veriliş: {t_odunc}")
            self.lbl_teslim_tarih_son.setText(f"⏳ Son Teslim: {t_son}")
                
        else:
            self.secili_odunc_id = None
            self.lbl_teslim_uye.setText("Seçim Bekleniyor...")
            self.lbl_teslim_kitap.setText("-")
            self.lbl_teslim_tarih_verilis.setText("-")
            self.lbl_teslim_tarih_son.setText("-")

    def kitap_teslim_al(self):
        if not hasattr(self, 'secili_odunc_id') or self.secili_odunc_id is None:
            QMessageBox.warning(self, "Uyarı", "Lütfen listeden teslim alınacak kaydı seçin.")
            return

        if QMessageBox.question(self, "Onay", "Kitabı teslim almak istiyor musunuz?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return

        bugun = datetime.now().strftime('%Y-%m-%d')
        basari, mesaj, ceza_var = self.db.kitap_teslim_al_prosedur(self.secili_odunc_id, bugun)
        
        if basari:
            if ceza_var:
                QMessageBox.warning(self, "Teslim Alındı (GECİKMELİ)", f"{mesaj}")
            else:
                QMessageBox.information(self, "Başarılı", "Kitap zamanında teslim alındı.")
            
            self.teslim_listesi_yukle()
            self.kitap_listele()
            self.odunc_kitap_listele()
            self.secili_odunc_id = None
            self.lbl_teslim_uye.setText("-")
            self.lbl_teslim_kitap.setText("-")
            
            
        else:
            QMessageBox.critical(self, "Hata", f"İşlem sırasında hata oluştu:\n{mesaj}")
            
            
    
    #           CEZA EKRANI FONKSİYONLARI

    def ceza_ekrani_hazirlik(self):
        from PyQt5.QtCore import QDate
        bugun = QDate.currentDate()
        self.date_ceza_baslangic.setDate(bugun.addMonths(-1))
        self.date_ceza_bitis.setDate(bugun)

        self.cmb_ceza_uye.clear()
        self.cmb_ceza_uye.addItem("Tüm Üyeler", -1) 
            
        uyeler = self.db.ceza_icin_uyeleri_getir()
        for u in uyeler:
            self.cmb_ceza_uye.addItem(f"{u[1]} {u[2]}", u[0])

    def cezalari_listele(self):
        secili_uye_index = self.cmb_ceza_uye.currentIndex()
        uye_id = self.cmb_ceza_uye.itemData(secili_uye_index)
        
        baslangic = self.date_ceza_baslangic.date().toString("yyyy-MM-dd")
        bitis = self.date_ceza_bitis.date().toString("yyyy-MM-dd")

        cezalar = self.db.cezalar_filtreli(uye_id, baslangic, bitis)

        self.tbl_cezalar.setRowCount(0)
        self.tbl_cezalar.setRowCount(len(cezalar))
        
        toplam_borc = 0.0

        for satir, veri in enumerate(cezalar):
            self.tbl_cezalar.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_cezalar.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            self.tbl_cezalar.setItem(satir, 2, QTableWidgetItem(str(veri[2])))
            self.tbl_cezalar.setItem(satir, 3, QTableWidgetItem(f"{veri[3]} TL"))
            self.tbl_cezalar.setItem(satir, 4, QTableWidgetItem(str(veri[4])))
            self.tbl_cezalar.setItem(satir, 5, QTableWidgetItem(str(veri[5])))
            
            try:
                toplam_borc += float(veri[3])
            except:
                pass

        self.lbl_ceza_toplam.setText(f"Toplam: {toplam_borc:.2f} TL")
   
            
    #           RAPORLAMA FONKSİYONLARI

    def rapor_ekrani_hazirlik(self):
        from PyQt5.QtCore import QDate
        bugun = QDate.currentDate()
        
        self.date_rapor_baslangic.setDate(bugun.addMonths(-1))
        self.date_rapor_bitis.setDate(bugun)

        self.cmb_rapor_uye.clear()
        self.cmb_rapor_uye.addItem("Tüm Üyeler", -1)
            
        uyeler = self.db.ceza_icin_uyeleri_getir()
        for u in uyeler:
            self.cmb_rapor_uye.addItem(f"{u[1]} {u[2]}", u[0])

    def rapor_hareket_listele(self):
        idx = self.cmb_rapor_uye.currentIndex()
        uye_id = self.cmb_rapor_uye.itemData(idx)
        
        baslangic = self.date_rapor_baslangic.date().toString("yyyy-MM-dd")
        bitis = self.date_rapor_bitis.date().toString("yyyy-MM-dd")
        
        sonuclar = self.db.rapor_odunc_hareketleri(baslangic, bitis, uye_id)
        
        self.tbl_rapor_hareket.setRowCount(0)
        self.tbl_rapor_hareket.setRowCount(len(sonuclar))
        
        for satir, veri in enumerate(sonuclar):
            self.tbl_rapor_hareket.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_rapor_hareket.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            self.tbl_rapor_hareket.setItem(satir, 2, QTableWidgetItem(str(veri[2])))
            self.tbl_rapor_hareket.setItem(satir, 3, QTableWidgetItem(str(veri[3])))
            self.tbl_rapor_hareket.setItem(satir, 4, QTableWidgetItem(str(veri[5]))) 
    def rapor_geciken_listele(self):
        gecikenler = self.db.rapor_gecikenler()
        
        self.tbl_rapor_geciken.setRowCount(0)
        self.tbl_rapor_geciken.setRowCount(len(gecikenler))
        
        for satir, veri in enumerate(gecikenler):
            self.tbl_rapor_geciken.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_rapor_geciken.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            self.tbl_rapor_geciken.setItem(satir, 2, QTableWidgetItem(str(veri[2])))
            self.tbl_rapor_geciken.setItem(satir, 3, QTableWidgetItem(str(veri[3])))
            
            item_gecikme = QTableWidgetItem(f"{veri[4]} Gün")
            item_gecikme.setForeground(Qt.red) # Kırmızı yazı
            self.tbl_rapor_geciken.setItem(satir, 4, item_gecikme)

    def rapor_populer_listele(self):
        populerler = self.db.rapor_populer_kitaplar(limit=10) # Top 10
        
        self.tbl_rapor_populer.setRowCount(0)
        self.tbl_rapor_populer.setRowCount(len(populerler))
        
        for satir, veri in enumerate(populerler):
            self.tbl_rapor_populer.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_rapor_populer.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            self.tbl_rapor_populer.setItem(satir, 2, QTableWidgetItem(str(veri[2])))        
    
    def rapor_aktif_listele(self):
        veriler = self.db.rapor_en_aktif_uyeler(limit=10)
        
        self.tbl_rapor_aktif.setRowCount(0)
        self.tbl_rapor_aktif.setRowCount(len(veriler))
        
        for satir, veri in enumerate(veriler):
            self.tbl_rapor_aktif.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_rapor_aktif.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            
            item_sayi = QTableWidgetItem(str(veri[2]))
            item_sayi.setTextAlignment(Qt.AlignCenter)
            self.tbl_rapor_aktif.setItem(satir, 2, item_sayi)

    def rapor_ceza_listele(self):
        veriler = self.db.rapor_en_cok_ceza_alanlar(limit=10)
        
        self.tbl_rapor_ceza_istatistik.setRowCount(0)
        self.tbl_rapor_ceza_istatistik.setRowCount(len(veriler))
        
        for satir, veri in enumerate(veriler):
            self.tbl_rapor_ceza_istatistik.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_rapor_ceza_istatistik.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            
            item_tutar = QTableWidgetItem(f"{veri[2]} TL")
            item_tutar.setForeground(Qt.red)
            self.tbl_rapor_ceza_istatistik.setItem(satir, 2, item_tutar)        

    #      DİNAMİK SORGU FONKSİYONLARI

    def sorgu_ekrani_hazirlik(self):
        self.cmb_sorgu_kategori.clear()
        self.cmb_sorgu_kategori.addItem("Tüm Kategoriler", -1)
            
        kats = self.db.kategorileri_getir()
        for k in kats:
            self.cmb_sorgu_kategori.addItem(k[1], k[0]) 

    def dinamik_sorgu_calistir(self):
        filtreler = {}
        
        filtreler['ad'] = self.txt_sorgu_kitap.text()
        filtreler['yazar'] = self.txt_sorgu_yazar.text()
        
        idx = self.cmb_sorgu_kategori.currentIndex()
        filtreler['kategori_id'] = self.cmb_sorgu_kategori.itemData(idx)
            
        val = self.spin_sorgu_min.value()
        if val > 0: filtreler['yil_min'] = val
            
        val = self.spin_sorgu_max.value()
        if val > 0: filtreler['yil_max'] = val
            
        filtreler['sadece_mevcut'] = self.chk_sorgu_stok.isChecked()
            
        ham_veri = self.cmb_sorgu_sirala.currentText()
        temiz_veri = ham_veri.replace(" ", "").replace("ı", "I").replace("i", "İ").upper()
        filtreler['sirala_alan'] = temiz_veri
       
        secim = self.cmb_sorgu_yon.currentText()
        filtreler['sirala_yon'] = "DESC" if "Azalan" in secim else "ASC"

        sonuclar = self.db.dinamik_kitap_sorgula(filtreler)
        
        self.tbl_sorgu_sonuc.setRowCount(0)
        self.tbl_sorgu_sonuc.setRowCount(len(sonuclar))
        
        self.tbl_sorgu_sonuc.setColumnCount(4) 
        for satir, veri in enumerate(sonuclar):
                
            self.tbl_sorgu_sonuc.setItem(satir, 0, QTableWidgetItem(str(veri[0])))
            self.tbl_sorgu_sonuc.setItem(satir, 1, QTableWidgetItem(str(veri[1])))
            self.tbl_sorgu_sonuc.setItem(satir, 2, QTableWidgetItem(str(veri[2])))
            self.tbl_sorgu_sonuc.setItem(satir, 3, QTableWidgetItem(str(veri[3])))
                
        
        adet = len(sonuclar)
        self.lbl_sonuc_sayisi.setText(f"Toplam {adet} kayıt bulundu.")        
                
                
    def sonuclari_excele_aktar(self):
        row_count = self.tbl_sorgu_sonuc.rowCount()
        col_count = self.tbl_sorgu_sonuc.columnCount()
        
        if row_count == 0:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek veri yok!")
            return

        # Dosya uzantısını xlsx yaptık
        dosya_yolu, _ = QFileDialog.getSaveFileName(self, "Excel Olarak Kaydet", "", "Excel Dosyası (*.xlsx)")
        
        if dosya_yolu:
            if not dosya_yolu.endswith(".xlsx"):
                dosya_yolu += ".xlsx"

            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Kitap Sorgu Sonuçları"

                basliklar = []
                for col in range(col_count):
                    baslik = self.tbl_sorgu_sonuc.horizontalHeaderItem(col).text()
                    sheet.cell(row=1, column=col+1, value=baslik) 
                    basliklar.append(baslik)

                for row in range(row_count):
                    for col in range(col_count):
                        item = self.tbl_sorgu_sonuc.item(row, col)
                        veri = item.text() if item else ""
                        
                        try:
                            if veri.isdigit():
                                veri = int(veri)
                        except:
                            pass
                            
                        sheet.cell(row=row+2, column=col+1, value=veri) 

                workbook.save(dosya_yolu)
                QMessageBox.information(self, "Başarılı", "Excel dosyası başarıyla oluşturuldu!")
                
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu:\n{e}")
    def filtreleri_temizle(self):
        
        self.txt_sorgu_kitap.clear()
        self.txt_sorgu_yazar.clear()
        
        self.cmb_sorgu_kategori.setCurrentIndex(0)
        self.cmb_sorgu_sirala.setCurrentIndex(0)
        self.cmb_sorgu_yon.setCurrentIndex(0)

        self.spin_sorgu_min.setValue(0)
        self.spin_sorgu_max.setValue(0)
        
        self.chk_sorgu_stok.setChecked(False)
        
        self.tbl_sorgu_sonuc.setRowCount(0)
        self.lbl_sonuc_sayisi.setText("Filtreler temizlendi.")            
  
    def sorgu_sayfasini_ac(self):
        self.sorgu_ekrani_hazirlik()
        
        self.stackedWidget.setCurrentIndex(6)             
               
class GirisEkrani(QMainWindow):
    def __init__(self):
        super().__init__()
        
        try:
            uic.loadUi('giris.ui', self)
        except FileNotFoundError:
            print("HATA: giris.ui dosyasi bulunamadi! Dosya ismini kontrol et.")
            return
    
        self.btn_giris.clicked.connect(self.giris_yap)
        self.txt_sifre.returnPressed.connect(self.giris_yap)
        self.txt_kadi.returnPressed.connect(self.giris_yap)
        
        self.txt_kadi.installEventFilter(self)
        self.txt_sifre.installEventFilter(self)
        self.btn_giris.installEventFilter(self)
        
        self.ekrani_ortala()
    
    def eventFilter(self, source, event):
        if event.type() == QEvent.KeyPress:
            
            if event.key() == Qt.Key_Down:
                if source == self.txt_kadi:
                    self.txt_sifre.setFocus() 
                elif source == self.txt_sifre:
                    self.btn_giris.setFocus() 
            
            elif event.key() == Qt.Key_Up:
                if source == self.txt_sifre:
                    self.txt_kadi.setFocus()  
                elif source == self.btn_giris:
                    self.txt_sifre.setFocus() 
                    
        return super().eventFilter(source, event)
       
    def ekrani_ortala(self):
        qr = self.frameGeometry()
        cp = QApplication.desktop().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        
    def giris_yap(self):
        kadi = self.txt_kadi.text()
        sifre = self.txt_sifre.text()

        if not kadi or not sifre:
            QMessageBox.warning(self, "Uyari", "Kullanici adi ve sifre bos olamaz.")
            return

        db = Veritabani()
        kullanici = db.kullanici_dogrula(kadi, sifre)

        if kullanici:
            self.hide() 
            self.ana_sayfa = AnaSayfa(kullanici) 
            self.ana_sayfa.show()
        else:
            QMessageBox.critical(self, "Hata", "Kullanici adi veya sifre hatali!")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    pencere = GirisEkrani()
    pencere.show()
    sys.exit(app.exec_())
    